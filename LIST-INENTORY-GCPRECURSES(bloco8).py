import datetime
from google.cloud import asset_v1
from openpyxl import Workbook

# Inicializa o cliente do Asset
client = asset_v1.AssetServiceClient()


# Lista de projetos
projects = [
    #'shared-project-fca',
    #'non-prod-project-fca',
    #'prod-project-fca',
    #'sap-shared-sa',
    #'sap-nonprod-sa',
    #'sap-prod-sa',
    'banco-fidis-prod',
    'banco-fidis-nonprod'
    #fse-lab-custom',
    #'fse-lav-finance',
    #'fse-lab-infra',
    #'fse-lab-payroll',
    #'fse-lab-webapplication',
    #'fse-network',
    #'fse-non-prod',
    #'fse-prod',
]
# Cria um objeto Workbook
workbook = Workbook()

# Loop pelos projetos
for project_id in projects:
    print(f"Coletando recursos para o projeto: {project_id}")

    # Cria uma nova guia com o nome do projeto
    sheet = workbook.create_sheet(project_id)
    sheet.append(['Service', 'Resource'])

    # Obtém os recursos do projeto
    response = client.search_all_resources(
        scope=f"projects/{project_id}"
    )

    # Lista os recursos
    resources = list(response)

    # Verifica se existem recursos
    if resources:
        # Adiciona os nomes dos recursos na planilha
        for resource in resources:
            asset_type = resource.asset_type.split('/')[-1]  # Obtém apenas o nome do serviço
            resource_name = resource.name.replace('//', '/')  # Remove as barras duplas

            sheet.append([asset_type, resource_name])

# Remove a guia padrão criada automaticamente
workbook.remove(workbook['Sheet'])

# Cria uma nova guia "TotalServices"
sheet_total_services = workbook.create_sheet("Service Counts")
sheet_total_services.append(['Service', 'Total'])

# Dicionário para contar o total de serviços
service_counts = {}

# Loop pelos projetos
for project_id in projects:
    # Obtém os recursos do projeto
    response = client.search_all_resources(
        scope=f"projects/{project_id}"
    )

    # Lista os recursos
    resources = list(response)

    # Conta o número de serviços
    for resource in resources:
        asset_type = resource.asset_type.split('/')[-1]  # Obtém apenas o nome do serviço
        if asset_type in service_counts:
            service_counts[asset_type] += 1
        else:
            service_counts[asset_type] = 1

# Adiciona o total de serviços na guia "TotalServices"
for service, count in service_counts.items():
    sheet_total_services.append([service, count])

# Salva o arquivo Excel
workbook.save('gcp_inventory.xlsx')

# Imprime a mensagem de finalização
print("O programa foi finalizado. O arquivo Excel foi gerado com sucesso.")
