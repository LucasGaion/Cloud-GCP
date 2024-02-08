import datetime
from google.cloud import asset_v1
from openpyxl import Workbook

# Inicializa o cliente do Asset
client = asset_v1.AssetServiceClient()

# Lista de projetos
projects = [
    #'shared-project-fca',
    #'non-prod-project-fca',
    #'prod-project-fca',
    #'sap-shared-sa',
    #sap-nonprod-sa',
    #'sap-prod-sa',

    #'banco-fidis-prod',
    'banco-fidis-nonprod'
]


# Função para coletar os recursos de um projeto
def collect_resources(project_id):
    resources = []
    response = client.search_all_resources(scope=f"projects/{project_id}")
    resources.extend(list(response))
    return resources

# Cria um objeto Workbook
workbook = Workbook()

# Dicionário para contar o total de serviços
service_counts = {}

# Loop pelos projetos
for project_id in projects:
    print(f"Coletando recursos para o projeto: {project_id}")
    # Obtém os recursos do projeto
    resources = collect_resources(project_id)
    print(f"Total de recursos coletados para {project_id}: {len(resources)}")

    # Cria uma nova guia com o nome do projeto
    sheet = workbook.create_sheet(project_id)
    sheet.append(['Serviço', 'Recurso'])

    # Adiciona os nomes dos recursos na planilha e conta o número de serviços
    for resource in resources:
        asset_type = resource.asset_type.split('/')[-1]  # Obtém apenas o nome do serviço
        resource_name = resource.name.replace('//', '/')  # Remove as barras duplas
        sheet.append([asset_type, resource_name])

        # Conta o número de serviços
        if asset_type in service_counts:
            service_counts[asset_type] += 1
        else:
            service_counts[asset_type] = 1

# Remove a guia padrão criada automaticamente
workbook.remove(workbook['Sheet'])

# Adiciona o total de serviços em outra guia
sheet_total_services = workbook.create_sheet("Service Counts")
sheet_total_services.append(['Service', 'Total'])

# Adiciona o total de serviços na guia "Service Counts"
for service, count in service_counts.items():
    sheet_total_services.append([service, count])

# Salva o arquivo Excel
workbook.save('banco-fidis-nonprod.xlsx')

# Imprime a mensagem de finalização
print("O programa foi finalizado. O arquivo Excel foi gerado com sucesso.")
