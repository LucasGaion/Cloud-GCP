import datetime
from google.cloud import asset_v1
from openpyxl import Workbook

# Inicializa o cliente do Asset
client = asset_v1.AssetServiceClient()

# Lista de projetos
projects = [
    #'shared-project-fca',
    'non-prod-project-fca',
    #'prod-project-fca',
    #'sap-shared-sa',
    #'sap-nonprod-sa',
    #'sap-prod-sa',
]

# Lista de serviços do Google Cloud Platform (GCP)
services = [
    'compute.googleapis.com/instance/disk/write_bytes_count'
]

# Função para coletar os recursos de um projeto
def collect_resources(project_id):
    resources = []
    response = client.search_all_resources(scope=f"projects/{project_id}")
    resources.extend(list(response))
    return resources

# Filtrar os recursos para incluir apenas os serviços especificados
def filter_resources(resources):
    filtered_resources = []
    for resource in resources:
        asset_type = resource.asset_type
        if any(service in asset_type for service in services):
            filtered_resources.append(resource)
    return filtered_resources

# Cria um objeto Workbook
workbook = Workbook()

# Dicionário para contar o total de serviços
service_counts = {service: 0 for service in services}

# Loop pelos projetos
for project_id in projects:
    print(f"Coletando recursos para o projeto: {project_id}")
    # Obtém os recursos do projeto
    resources = collect_resources(project_id)
    print(f"Total de recursos coletados: {len(resources)}")

    # Cria uma nova guia com o nome do projeto
    sheet = workbook.create_sheet(project_id)
    sheet.append(['Serviço', 'Recurso'])

    # Filtra os recursos para incluir apenas os serviços especificados
    filtered_resources = filter_resources(resources)
    print(f"Total de recursos filtrados: {len(filtered_resources)}")

    # Adiciona os nomes dos recursos na planilha
    for resource in filtered_resources:
        asset_type = resource.asset_type.split('/')[-1]  # Obtém apenas o nome do serviço
        resource_name = resource.name.replace('//', '/')  # Remove as barras duplas
        sheet.append([asset_type, resource_name])
        print(resource.asset_type)

    # Conta o número de serviços
    for service in services:
        service_counts[service] += sum(1 for resource in filtered_resources if service in resource.asset_type)

# Remove a guia padrão criada automaticamente
workbook.remove(workbook['Sheet'])

# Cria uma nova guia "Total de Serviços"
sheet_total_services = workbook.create_sheet("Total de Serviços")
sheet_total_services.append(['Serviço', 'Total'])

# Adiciona o total de serviços na guia "Total de Serviços"
for service, count in service_counts.items():
    sheet_total_services.append([service, count])

# Verifica se algum serviço está ausente e adiciona com total igual a 0
for service in services:
    if service not in service_counts:
        sheet_total_services.append([service, 0])

# Salva o arquivo Excel
workbook.save('sap-shared-sa1.xlsx')

# Imprime a mensagem de finalização
print("O programa foi finalizado. O arquivo Excel foi gerado com sucesso.")
