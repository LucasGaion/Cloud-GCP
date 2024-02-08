import datetime
from google.cloud import asset_v1
from openpyxl import Workbook

# Inicializa o cliente do Asset
client = asset_v1.AssetServiceClient()

# Lista de projetos
projects = [
    'shared-project-fca',
    'non-prod-project-fca',
    'prod-project-fca',
    'sap-shared-sa',
    'sap-nonprod-sa',
    'sap-prod-sa',
    'banco-fidis-prod',
    'banco-fidis-nonprod'
]
# Lista de serviços do Google Cloud Platform (GCP)
services = [
    'compute.googleapis.com/BackendService', # load balancer
    'dns.googleapis.com/ManagedZone', # Dns
    'dns.googleapis.com/Policy',
    'aiplatform.googleapis.com/CustomJob',
    'aiplatform.googleapis.com/Featurestore',
    'aiplatform.googleapis.com/HyperparameterTuningJob',
    'aiplatform.googleapis.com/Job',
    'aiplatform.googleapis.com/MetadataStore',
    'aiplatform.googleapis.com/Model',
    'aiplatform.googleapis.com/PipelineJob',
    'aiplatform.googleapis.com/Version',
    'aiplatform.googleapis.com/Tensorboard',
    'appengine.googleapis.com/Application',
    'artifactregistry.googleapis.com/Repository',
    'bigquery.googleapis.com/Dataset',
    'bigquery.googleapis.com/Table',
    'bigtable.googleapis.com/Table',
    'billing.googleapis.com/ProjectBillingInfo',
    'cloudfunctions.googleapis.com/CloudFunction',
    'cloudfunctions.googleapis.com/Function',
    'cloudbuild.googleapis.com/Trigger',
    'cloudkms.googleapis.com/CryptoKey',
    'cloudkms.googleapis.com/CryptoKeyVersion',
    'cloudkms.googleapis.com/KeyRing',
    'cloudresourcemanager.googleapis.com/Folder',
    'cloudresourcemanager.googleapis.com/Organization',
    'cloudresourcemanager.googleapis.com/Project',
    'cloudtasks.googleapis.com/Queue',
    'composer.googleapis.com/Connector',
    'composer.googleapis.com/Workflow',
    'compute.googleapis.com/Address',
    'compute.googleapis.com/Autoscaler',
    'compute.googleapis.com/BackendBucket',
    'compute.googleapis.com/BatchPredictionJob',
    'compute.googleapis.com/Commitment',
    'compute.googleapis.com/Database',
    'compute.googleapis.com/Deployment',
    'compute.googleapis.com/Disks',
    'compute.googleapis.com/Disk',
    'compute.googleapis.com/DockerImage',
    'compute.googleapis.com/Environment',
    'compute.googleapis.com/Firewall',
    'compute.googleapis.com/ForwardingRule',
    'compute.googleapis.com/GlobalForwardingRule',
    'compute.googleapis.com/HealthCheck',
    'compute.googleapis.com/HttpHealthCheck',
    'compute.googleapis.com/Image',
    'compute.googleapis.com/Instance',
    'compute.googleapis.com/InstanceGroup',
    'compute.googleapis.com/InstanceGroupManager',
    'compute.googleapis.com/InstanceTemplate',
    'compute.googleapis.com/Key',
    'compute.googleapis.com/License',
    'compute.googleapis.com/LoadBalancer',
    'compute.googleapis.com/LogBucket',
    'compute.googleapis.com/MavenArtifact',
    'compute.googleapis.com/ManagedService',
    'compute.googleapis.com/Network',
    'compute.googleapis.com/NetworkEndpointGroup',
    'compute.googleapis.com/Node',
    'compute.googleapis.com/NodePool',
    'compute.googleapis.com/Notification',
    'compute.googleapis.com/Object',
    'compute.googleapis.com/Project',
    'compute.googleapis.com/Pod',
    'compute.googleapis.com/ReplicaSet',
    'compute.googleapis.com/Repository',
    'compute.googleapis.com/ResourcePolicy',
    'compute.googleapis.com/Revision',
    'compute.googleapis.com/Route',
    'compute.googleapis.com/Router',
    'compute.googleapis.com/Role',
    'compute.googleapis.com/RoleBinding',
    'compute.googleapis.com/Snapshot',
    'compute.googleapis.com/SslCertificate',
    'compute.googleapis.com/SslPolicy',
    'compute.googleapis.com/SecurityPolicy',
    'compute.googleapis.com/SecurityPolicy',
    'compute.googleapis.com/Service',
    'compute.googleapis.com/Subnetwork',
    'compute.googleapis.com/TargetHttpProxy',
    'compute.googleapis.com/TargetHttpsProxy',
    'compute.googleapis.com/TargetTcpProxy',
    'compute.googleapis.com/TargetSslProxy',
    'compute.googleapis.com/TargetVpnGateway',
    'compute.googleapis.com/TargetPool',
    'compute.googleapis.com/TargetHttpsProxy',
    'compute.googleapis.com/UrlMap',
    'compute.googleapis.com/VpnTunnel',
    'container.googleapis.com/Cluster',
    'container.googleapis.com/DockerImage',
    'container.googleapis.com/NodePool',
    'dataflow.googleapis.com/Job',
    'dataflow.googleapis.com/Snapshot',
    'dataproc.googleapis.com/Cluster',
    'dataproc.googleapis.com/Job',
    'datastore.googleapis.com/Backup',
    'datastore.googleapis.com/Index',
    'datastore.googleapis.com/Kind',
    'datastore.googleapis.com/Namespace',
    'deploymentmanager.googleapis.com/Deployment',
    'endpoints.googleapis.com/Api',
    'endpoints.googleapis.com/ApiConfig',
    'endpoints.googleapis.com/Gateway',
    'endpoints.googleapis.com/Service',
    'firebase.googleapis.com/FirebaseAppInfo',
    'firebase.googleapis.com/FirebaseProject',
    'firebaserules.googleapis.com/Release',
    'firebaserules.googleapis.com/Ruleset',
    'firestore.googleapis.com/Document',
    'firestore.googleapis.com/Index',
    'firestore.googleapis.com/Project',
    'firestore.googleapis.com/Schema',
    'firestore.googleapis.com/Target',
    'gce-api.googleapis.com/Workspace',
    'iam.googleapis.com/ServiceAccount',
    'iam.googleapis.com/ServiceAccountKey',
    'kms.googleapis.com/CryptoKey',
    'kms.googleapis.com/CryptoKeyVersion',
    'logging.googleapis.com/BillingAccount',
    'logging.googleapis.com/Folder',
    'logging.googleapis.com/Organization',
    'logging.googleapis.com/Project',
    'logging.googleapis.com/Sink',
    'ml.googleapis.com/Model',
    'ml.googleapis.com/Version',
    'monitoring.googleapis.com/AlertPolicy',
    'networking.gke.io/NetworkPolicy',
    'notebooks.googleapis.com/Environment',
    'pubsub.googleapis.com/Project',
    'pubsub.googleapis.com/Subscription',
    'pubsub.googleapis.com/Topic',
    'rbac.authorization.k8s.io/ClusterRole',
    'rbac.authorization.k8s.io/ClusterRoleBinding',
    'rbac.authorization.k8s.io/Role',
    'rbac.authorization.k8s.io/RoleBinding',
    'secretmanager.googleapis.com/Secret',
    'secretmanager.googleapis.com/SecretVersion',
    'securitycenter.googleapis.com/LocationSettings',
    'sqladmin.googleapis.com/Database',
    'storage.googleapis.com/Bucket',
    'storage.googleapis.com/HmacKey',
    'storage.googleapis.com/Notification',
    'storage.googleapis.com/Object',
    'storage.googleapis.com/Project',
    'composer.googleapis.com/Service',
    'container.googleapis.com/ClusterAutoscaler',
    'container.googleapis.com/Operation',
    'container.googleapis.com/PodSecurityPolicy',
    'container.googleapis.com/Cluster',
    'container.googleapis.com/DockerImage',
    'container.googleapis.com/NodePool',
    'container.googleapis.com/Settings',
    'container.googleapis.com/ClusterTelemetry',
    'dataproc.googleapis.com/WorkflowTemplate',
    'file.googleapis.com/FileSystem',
    'iap.googleapis.com/IdentityAwareProxyClient',
    'iam.googleapis.com/ExternalCredential',
    'iam.googleapis.com/WorkloadIdentityPool',
    'managedidentities.googleapis.com/Domain',
    'memcache.googleapis.com/Instance',
    'monitoring.googleapis.com/MetricDescriptor',
    'monitoring.googleapis.com/MonitoredProject',
    'monitoring.googleapis.com/UptimeCheckConfig',
    'notebooks.googleapis.com/Instance',
    'notebooks.googleapis.com/InstanceConfig',
    'notebooks.googleapis.com/Runtime',
    'notebooks.googleapis.com/Schedule',
    'notebooks.googleapis.com/Environment',
    'notebooks.googleapis.com/EnvironmentConfig',
    'notebooks.googleapis.com/Location',
    'notebooks.googleapis.com/ManagedNotebook',
    'notebooks.googleapis.com/ManagedNotebookConfig',
    'notebooks.googleapis.com/ManagedNotebookService',
    'notebooks.googleapis.com/VmImage',
    'osconfig.googleapis.com/InstanceOSPoliciesCompliance',
    'osconfig.googleapis.com/OSPolicyAssignment',
    'osconfig.googleapis.com/OSPolicy',
    'osconfig.googleapis.com/Instance',
    'osconfig.googleapis.com/GuestPolicy',
    'oslogin.googleapis.com/Profile',
    'oslogin.googleapis.com/SshPublicKey',
    'oslogin.googleapis.com/PosixAccount',
    'privateca.googleapis.com/CertificateAuthority',
    'privateca.googleapis.com/Certificate',
    'privateca.googleapis.com/CertificateRevocationList',
    'privateca.googleapis.com/CertificateTemplate',
    'run.googleapis.com/Configuration',
    'run.googleapis.com/DomainMapping',
    'run.googleapis.com/Revision',
    'run.googleapis.com/Route',
    'run.googleapis.com/Service',
    'servicedirectory.googleapis.com/Namespace',
    'servicedirectory.googleapis.com/Service',
    'servicemanagement.googleapis.com/Service',
    'servicenetworking.googleapis.com/Connection',
    'servicenetworking.googleapis.com/ConsumerQuotaMetric',
    'servicenetworking.googleapis.com/ConsumerQuotaLimit',
    'servicenetworking.googleapis.com/ConsumerQuota',
    'servicenetworking.googleapis.com/Service',
    'servicenetworking.googleapis.com/Visibility',
    'vpcaccess.googleapis.com/Connector',
    'vpcaccess.googleapis.com/Operation',
    'networkmanagement.googleapis.com/ConnectivityTest',
    'compute.googleapis.com/InterconnectAttachment',
    'backup.googleapis.com/BackupRun',
    'compute.googleapis.com/NetworkEndpointGroups',
    'logging.googleapis.com/LogSink',
    'serviceusage.googleapis.com/Services',
    'servicedirectory.googleapis.com/Namespace',
]

# Função para coletar os recursos de um projeto
def collect_resources(project_id):
    resources = []
    response = client.search_all_resources(scope=f"projects/{project_id}")
    resources.extend(list(response))
    return resources

# Filtrar os recursos para incluir apenas os serviços especificados
def filter_resources(resources):
    filtered_resources = []
    for resource in resources:
        asset_type = resource.asset_type
        if any(service in asset_type for service in services):
            filtered_resources.append(resource)
    return filtered_resources

# Cria um objeto Workbook
workbook = Workbook()

# Dicionário para contar o total de serviços
service_counts = {service: 0 for service in services}

# Loop pelos projetos
for project_id in projects:
    print(f"Coletando recursos para o projeto: {project_id}")
    # Obtém os recursos do projeto
    resources = collect_resources(project_id)
    print(f"Total de recursos coletados: {len(resources)}")

    # Cria uma nova guia com o nome do projeto
    sheet = workbook.create_sheet(project_id)
    sheet.append(['Serviço', 'Recurso'])

    # Filtra os recursos para incluir apenas os serviços especificados
    filtered_resources = filter_resources(resources)
    print(f"Total de recursos filtrados: {len(filtered_resources)}")

    # Adiciona os nomes dos recursos na planilha
    for resource in filtered_resources:
        asset_type = resource.asset_type.split('/')[-1]  # Obtém apenas o nome do serviço
        resource_name = resource.name.replace('//', '/')  # Remove as barras duplas
        sheet.append([asset_type, resource_name])
        print(resource.asset_type)

    # Conta o número de serviços
    for service in services:
        service_counts[service] += sum(1 for resource in filtered_resources if service in resource.asset_type)

# Remove a guia padrão criada automaticamente
workbook.remove(workbook['Sheet'])

# Cria uma nova guia "Total de Serviços"
sheet_total_services = workbook.create_sheet("Total de Serviços")
sheet_total_services.append(['Serviço', 'Total'])

# Adiciona o total de serviços na guia "Total de Serviços"
for service, count in service_counts.items():
    sheet_total_services.append([service, count])

# Verifica se algum serviço está ausente e adiciona com total igual a 0
for service in services:
    if service not in service_counts:
        sheet_total_services.append([service, 0])

# Salva o arquivo Excel
workbook.save('gcp_inventory5.xlsx')

# Imprime a mensagem de finalização
print("O programa foi finalizado. O arquivo Excel foi gerado com sucesso.")
